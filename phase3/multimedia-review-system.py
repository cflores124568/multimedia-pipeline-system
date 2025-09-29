import argparse
import sys
import os
import re
import csv
import json
import subprocess
from datetime import datetime
from pymongo import MongoClient
import getpass
import openpyxl
from openpyxl.drawing import image as xl_image
import requests
from PIL import Image
import tempfile
import shutil
from frameioclient import FrameioClient
from dotenv import load_dotenv

#Load environment variables
load_dotenv()

def find_file(filename):
    #Check if filename is full or relative path
    if os.path.isfile(filename):
        return filename
    #Search common directories if file isn't in current location
    search_paths = [
        '.',  #Current directory
        './input',  #Structured github folders
        './input/baselight',
        './input/flame',
        './input/xytech',
        os.path.expanduser('~/Downloads'),
        os.path.expanduser('~/Desktop'),
        os.path.expanduser('~/Documents'),
    ]
    for path in search_paths:
        full_path = os.path.join(path, filename)
        if os.path.isfile(full_path):
            return full_path
    return None

def is_valid_frame_data(data):
    #Reject invalid frame data like <null> or <err>
    if not data or not isinstance(data, str):
        return False
    data_str = data.strip().lower()
    invalid_values = ['<null>', '<err>', 'null', 'err', '']
    if data_str in invalid_values:
        return False
    return data_str.isdigit()

def parse_arguments():
    parser = argparse.ArgumentParser(description='Process Baselight/Flame and Xytech files with video processing for CSV, MongoDB, or Excel output')
    parser.add_argument("--files", nargs='+', dest="workFiles", help="List of Baselight/Flame files to process")
    parser.add_argument("--x", "--xytech", help="Path to the Xytech workorder file")
    parser.add_argument("--verbose", action="store_true", help="Print detailed processing information")
    parser.add_argument("--output", "--o", choices=['CSV', 'DB', 'XLS'], help="Choose output format: CSV, MongoDB, or Excel")
    parser.add_argument("--query", action="store_true", help="Run database queries")
    parser.add_argument("--view", action="store_true", help="Display contents of the database")
    parser.add_argument("--date", help="Export CSV for a specific date (format: YYYYMMDD)")
    parser.add_argument("--clear", action="store_true", help="Delete all database records (use with caution)")
    parser.add_argument("--process", help="Process video file")
    return parser.parse_args()

def parse_db_files(args):
    parsed_data = []
    xytech_data = None
    
    print("Processing input files...")
    if args.xytech:
        xytech_file = find_file(args.xytech)
        if not xytech_file:
            print(f"Error: Xytech file '{args.xytech}' not found")
            sys.exit(1)
        xytech_data = parse_xytech_file(xytech_file)
        if args.verbose:
            print(f"Found Xytech file: {xytech_file}")
            print(f"Parsed Xytech file: {xytech_data['job']} by {xytech_data['operator']}")
    
    for file_path in args.workFiles or []:
        found_file = find_file(file_path)
        if not found_file:
            print(f"Error: File '{file_path}' not found")
            continue
            
        if args.verbose:
            print(f"Found file: {found_file}")
            print(f"Processing file: {found_file}")
        
        filename_info = parse_filename(found_file)
        
        try:
            with open(found_file, 'r') as file:
                content = file.read().strip()
                if is_baselight_file(content):
                    file_data = parse_baselight_content(content, filename_info, xytech_data, args)
                elif is_flame_file(content):
                    file_data = parse_flame_content(content, filename_info, xytech_data, args)
                else:
                    print(f"Warning: Unknown file format for {found_file}")
                    continue
                
                parsed_data.extend(file_data)
                
        except FileNotFoundError:
            print(f"Error: File {found_file} not found")
        except Exception as e:
            print(f"Error processing {found_file}: {e}")
    
    print(f"Processed {len(args.workFiles or [])} files, parsed {len(parsed_data)} entries")
    return parsed_data, xytech_data

def parse_xytech_file(xytech_path):
    with open(xytech_path, 'r') as file:
        all_lines = [line.strip() for line in file if line.strip()]
    
    #Assuming specific line order for producer, operator, job
    producer = all_lines[1].split(":", 1)[1].strip() if len(all_lines) > 1 else ""
    operator = all_lines[2].split(":", 1)[1].strip() if len(all_lines) > 2 else ""
    job = all_lines[3].split(":", 1)[1].strip() if len(all_lines) > 3 else ""
    
    #Find location and notes sections dynamically
    location_start = all_lines.index("Location:") + 1 if "Location:" in all_lines else len(all_lines)
    notes_start = all_lines.index("Notes:") + 1 if "Notes:" in all_lines else len(all_lines)
    notes = all_lines[notes_start] if notes_start < len(all_lines) else ""
    location_list = all_lines[location_start:notes_start-1] if location_start < notes_start else []
    
    return {
        'producer': producer,
        'operator': operator,
        'job': job,
        'locations': location_list,
        'notes': notes
    }

def parse_filename(file_path):
    basename = os.path.basename(file_path)
    name_without_ext = os.path.splitext(basename)[0]
    
    parts = name_without_ext.split('_')
    if len(parts) != 3:
        raise ValueError(f"Filename {file_path} doesn't match expected format (machine_user_date)")
    
    machine, user, date_str = parts
    
    try:
        date_obj = datetime.strptime(date_str, '%Y%m%d')
    except ValueError:
        raise ValueError(f"Invalid date format in filename: {date_str}")
    
    return {
        'machine': machine,
        'user': user,
        'date': date_obj,
        'date_string': date_str,
        'filename': basename
    }

def is_baselight_file(content):
    lines = content.split('\n')
    for line in lines:
        line = line.strip()
        if line:
            parts = line.split()
            #Check for Baselight format: path followed by frame numbers
            if len(parts) >= 2 and parts[0].startswith('/') and not parts[0].startswith('/net/flame-archive'):
                try:
                    for part in parts[1:]:
                        int(part)
                    return True
                except ValueError:
                    continue
    return False

def is_flame_file(content):
    lines = content.split('\n')
    for line in lines:
        line = line.strip()
        #Flame files start with specific archive path
        if line and line.startswith('/net/flame-archive'):
            return True
    return False

PATH_KEYWORDS = ['production', 'baselightfilesystem1']

def get_logical_path(file_path):
    #Extract logical path starting from known keywords
    for keyword in PATH_KEYWORDS:
        if keyword in file_path:
            keyword_index = file_path.find(keyword)
            return file_path[keyword_index:]
    return file_path

def find_matching_xytech_location(baselight_flame_location, xytech_locations, args):
    if not xytech_locations:
        return baselight_flame_location
    
    bf_logical = get_logical_path(baselight_flame_location)
    
    if args.verbose:
        print(f"Looking for Xytech match for: '{bf_logical}'")
    
    #Case-insensitive path matching for consistency
    for xytech_location in xytech_locations:
        xytech_logical = get_logical_path(xytech_location.strip())
        if bf_logical.lower() == xytech_logical.lower():
            if args.verbose:
                print(f"  ✓ Found match: '{xytech_location.strip()}'")
            return xytech_location.strip()
    
    if args.verbose:
        print(f"  ✗ No match found, keeping original")
    
    return baselight_flame_location

def parse_baselight_content(content, filename_info, xytech_data, args):
    data = []
    lines = content.split('\n')
    location_frames = {}
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        parts = line.split()
        if len(parts) >= 2:
            original_path = parts[0]
            if original_path not in location_frames:
                location_frames[original_path] = []
            
            #Collect valid frame numbers for each path
            for part in parts[1:]:
                if not is_valid_frame_data(part):
                    if args.verbose:
                        print(f"Warning: Skipping invalid frame data '{part}' for location {original_path}")
                    continue
                try:
                    frame_number = int(part)
                    location_frames[original_path].append(frame_number)
                except ValueError:
                    if args.verbose:
                        print(f"Warning: Could not parse frame number '{part}' for location {original_path}")
                    continue
    
    for original_path in location_frames:
        unique_frames = list(set(location_frames[original_path]))
        frame_ranges = make_frame_ranges(unique_frames)
        final_location = find_matching_xytech_location(original_path, xytech_data['locations'], args) if xytech_data and xytech_data['locations'] else original_path
        
        entry = {
            'file_type': 'Baselight',
            'machine': filename_info['machine'],
            'user': filename_info['user'],
            'date': filename_info['date'],
            'filename': filename_info['filename'],
            'location': final_location,
            'frames': frame_ranges,
            'storage': original_path,
            'producer': xytech_data['producer'] if xytech_data else '',
            'operator': xytech_data['operator'] if xytech_data else '',
            'job': xytech_data['job'] if xytech_data else '',
            'notes': xytech_data['notes'] if xytech_data else ''
        }
        data.append(entry)
    
    return data

def parse_flame_content(content, filename_info, xytech_data, args):
    data = []
    lines = content.split('\n')
    location_frames = {}
    
    for line in lines:
        line = line.strip()
        if line:
            parts = line.split()
            if len(parts) >= 2 and parts[0].startswith('/net/flame-archive'):
                #Dynamically find where frame numbers start after path components
                frame_start_index = 1
                for i in range(1, len(parts)):
                    try:
                        int(parts[i])
                        frame_start_index = i
                        break
                    except ValueError:
                        continue
                
                location_parts = parts[1:frame_start_index]
                flame_location = ' '.join(location_parts)
                if flame_location not in location_frames:
                    location_frames[flame_location] = []
                
                for frame in parts[frame_start_index:]:
                    if not is_valid_frame_data(frame):
                        if args.verbose:
                            print(f"Warning: Skipping invalid frame data '{frame}' for location {flame_location}")
                        continue
                    try:
                        frame_number = int(frame)
                        location_frames[flame_location].append(frame_number)
                    except ValueError:
                        if args.verbose:
                            print(f"Warning: Could not parse frame number '{frame}' for location {flame_location}")
                        continue
    
    for flame_location in location_frames:
        unique_frames = list(set(location_frames[flame_location]))
        frame_ranges = make_frame_ranges(unique_frames)
        #Convert space-separated location to path format
        flame_path = flame_location.replace(' ', '/')
        if not flame_path.startswith('/'):
            flame_path = '/' + flame_path

        final_location = find_matching_xytech_location(flame_path, xytech_data['locations'], args) if xytech_data and xytech_data['locations'] else flame_path
        
        entry = {
            'file_type': 'Flame',
            'machine': filename_info['machine'],
            'user': filename_info['user'],
            'date': filename_info['date'],
            'filename': filename_info['filename'],
            'storage': '/net/flame-archive',
            'location': final_location,
            'frames': frame_ranges,
            'producer': xytech_data['producer'] if xytech_data else '',
            'operator': xytech_data['operator'] if xytech_data else '',
            'job': xytech_data['job'] if xytech_data else '',
            'notes': xytech_data['notes'] if xytech_data else ''
        }
        data.append(entry)
    
    return data

def clean_up_path(file_path):
    #Strip path to relevant portion after known keywords
    for keyword in PATH_KEYWORDS:
        if keyword in file_path:
            parts = file_path.split('/')
            for i in range(len(parts)):
                if parts[i] == keyword:
                    return '/'.join(parts[i+1:])
    return file_path

def make_frame_ranges(frame_numbers):
    if not frame_numbers:
        return []
    
    frame_numbers = sorted(frame_numbers)
    ranges = []
    start_frame = frame_numbers[0]
    end_frame = frame_numbers[0]
    
    #Group consecutive frames into ranges
    for frame in frame_numbers[1:]:
        if frame == end_frame + 1:
            end_frame = frame
        else:
            if start_frame == end_frame:
                ranges.append(str(start_frame))
            else:
                ranges.append(f"{start_frame}-{end_frame}")
            start_frame = frame
            end_frame = frame
    
    if start_frame == end_frame:
        ranges.append(str(start_frame))
    else:
        ranges.append(f"{start_frame}-{end_frame}")
    
    return ranges

def sort_frame_ranges_numeric(frame_ranges):
    #Sort frame ranges by their starting number
    def get_start_frame(frame_range):
        if '-' in frame_range:
            return int(frame_range.split('-')[0])
        else:
            return int(frame_range)
    return sorted(frame_ranges, key=get_start_frame)

def output_to_csv(parsed_data, xytech_data=None):
    if not parsed_data:
        print("No data to write to CSV")
        return
    
    output_dir = 'output'
    os.makedirs(output_dir, exist_ok=True)
    
    date_str = parsed_data[0]['date'].strftime('%Y%m%d') if parsed_data else datetime.now().strftime('%Y%m%d')
    output_file = os.path.join(output_dir, f'frame-fixes-{date_str}.csv')
    
    producer = xytech_data['producer'] if xytech_data else ''
    operator = xytech_data['operator'] if xytech_data else ''
    job = xytech_data['job'] if xytech_data else ''
    notes = xytech_data['notes'] if xytech_data else ''
    
    output_entries = []
    location_groups = {}
    for entry in parsed_data:
        location = entry['location']
        if location not in location_groups:
            location_groups[location] = []
        location_groups[location].extend(entry['frames'])
    
    #Prioritize Xytech locations, then add unmatched ones
    if xytech_data and xytech_data['locations']:
        for xytech_location in xytech_data['locations']:
            if xytech_location in location_groups:
                if location_groups[xytech_location]:
                    frame_ranges = sort_frame_ranges_numeric(location_groups[xytech_location])
                    for frame_range in frame_ranges:
                        output_entries.append([xytech_location, frame_range])
                else:
                    print(f"Warning: No frame data found for location: {xytech_location}")
                    output_entries.append([xytech_location, "No frames to fix"])
        
        for location in location_groups:
            if location not in xytech_data['locations']:
                print(f"Warning: Location {location} not found in Xytech file, adding at end of file")
                if location_groups[location]:
                    frame_ranges = sort_frame_ranges_numeric(location_groups[location])
                    for frame_range in frame_ranges:
                        output_entries.append([location, frame_range])
                else:
                    output_entries.append([location, "No frames to fix"])
    else:
        for location in location_groups:
            if location_groups[location]:
                frame_ranges = sort_frame_ranges_numeric(location_groups[location])
                for frame_range in frame_ranges:
                    output_entries.append([location, frame_range])
            else:
                print(f"Warning: No frame data found for location: {location}")
                output_entries.append([location, "No frames to fix"])
    
    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow([f"Producer: {producer}", f"Operator: {operator}", f"Job: {job}", f"Notes: {notes}"])
            writer.writerow([])
            writer.writerow([])
            writer.writerow(["Location:", "Frames to Fix:"])
            for entry in output_entries:
                writer.writerow(entry)
                
        print(f"Success! Output file created: {output_file}")
    except Exception as e:
        print(f"Error writing output file: {e}")
        sys.exit(1)

def get_video_info(video_path, args):
    try:
        cmd = [
            'ffprobe',
            '-v', 'quiet',
            '-print_format', 'json',
            '-show_format',
            '-show_streams',
            video_path
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            raise Exception(f"ffprobe error: {result.stderr}")
        
        data = json.loads(result.stdout)
        
        #Find the video stream to extract metadata
        video_stream = None
        for stream in data['streams']:
            if stream['codec_type'] == 'video':
                video_stream = stream
                break
        
        if not video_stream:
            raise Exception("No video stream found")
        
        duration = float(data['format']['duration'])
        fps_str = video_stream['r_frame_rate']
        #Handle fractional frame rates (e.g., 24/1)
        fps_parts = fps_str.split('/')
        fps = float(fps_parts[0]) / float(fps_parts[1]) if len(fps_parts) == 2 else float(fps_parts[0])
        total_frames = int(duration * fps)
        
        if args.verbose:
            print(f"Video duration: {duration:.2f} seconds")
            print(f"Frame rate: {fps:.2f} fps")
            print(f"Total frames: {total_frames}")
        
        return {
            'duration': duration,
            'fps': fps,
            'total_frames': total_frames,
            'width': video_stream.get('width', 0),
            'height': video_stream.get('height', 0)
        }
    
    except Exception as e:
        print(f"Error getting video info: {e}")
        return None

def frames_to_timecode(frame_number, fps):
    #Convert frame number to SMPTE timecode format
    total_seconds = frame_number / fps
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    seconds = int(total_seconds % 60)
    frames = int(frame_number % fps)
    
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}:{frames:02d}"

def parse_frame_range(frame_range_str):
    #Parse frame range string into start and end frames
    if '-' in frame_range_str:
        start_str, end_str = frame_range_str.split('-', 1)
        return int(start_str), int(end_str)
    else:
        frame_num = int(frame_range_str)
        return frame_num, frame_num

def get_matching_ranges(video_info, args):
    try:
        client = MongoClient('mongodb://localhost:27017/')
        db = client['MediaPipelineDB']
        files_collection = db['file_data']
        
        matching_ranges = []
        video_total_frames = video_info['total_frames']
        
        all_files = files_collection.find()
        
        #Clamp frame ranges to video duration
        for file_data in all_files:
            frames_list = file_data.get('frames', [])
            
            for frame_range_str in frames_list:
                start_frame, end_frame = parse_frame_range(frame_range_str)
                
                if start_frame <= video_total_frames and end_frame >= 1:
                    clamped_start = max(1, start_frame)
                    clamped_end = min(video_total_frames, end_frame)
                    
                    matching_range = file_data.copy()
                    matching_range['original_frame_range'] = frame_range_str
                    matching_range['clamped_start_frame'] = clamped_start
                    matching_range['clamped_end_frame'] = clamped_end
                    matching_range['start_timecode'] = frames_to_timecode(clamped_start, video_info['fps'])
                    matching_range['end_timecode'] = frames_to_timecode(clamped_end, video_info['fps'])
                    
                    middle_frame = (clamped_start + clamped_end) // 2
                    matching_range['middle_frame'] = middle_frame
                    matching_range['middle_timecode'] = frames_to_timecode(middle_frame, video_info['fps'])
                    
                    matching_ranges.append(matching_range)
        
        client.close()
        
        if args.verbose:
            print(f"Found {len(matching_ranges)} ranges within video bounds")
        
        return matching_ranges
    
    except Exception as e:
        print(f"Error querying database: {e}")
        return []

def create_thumbnail(video_path, frame_number, output_path, fps):
    try:
        time_seconds = frame_number / fps
        
        cmd = [
            'ffmpeg',
            '-y',
            '-i', video_path,
            '-ss', str(time_seconds),
            '-vframes', '1',
            '-s', '96x74',
            output_path
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            print(f"ffmpeg thumbnail error: {result.stderr}")
            return False
        
        return True
    
    except Exception as e:
        print(f"Error creating thumbnail: {e}")
        return False

def render_shot_segment(video_path, start_frame, end_frame, output_path, fps):
    try:
        start_time = start_frame / fps
        duration = (end_frame - start_frame + 1) / fps
        
        cmd = [
            'ffmpeg',
            '-y',
            '-i', video_path,
            '-ss', str(start_time),
            '-t', str(duration),
            '-c:v', 'libx264',
            '-crf', '23',
            output_path
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            print(f"ffmpeg render error: {result.stderr}")
            return False
        
        return True
    
    except Exception as e:
        print(f"Error rendering shot: {e}")
        return False

def upload_to_frameio(video_path, range_data=None):
    frameio_api_token = os.getenv("FRAMEIO_API_TOKEN")
    project_id = "https://next.frame.io/project/a0f47c90-485a-47a2-93cd-284b7bc0b3f9"
    root_asset_id = "a0f47c90-485a-47a2-93cd-284b7bc0b3f9"
    
    try:
        client = FrameioClient(frameio_api_token)
        asset = client.assets.upload(destination_id=root_asset_id, filepath=video_path)
        print(f"{video_path} has been uploaded to Frame.io")
        return True
    except Exception as e:
        print(f"Failed to upload {video_path} to Frame.io: {e}")
        return False

def output_to_excel(parsed_data, video_path=None, args=None):
    if not parsed_data:
        print("No data to write to Excel")
        return
    
    date_str = parsed_data[0]['date'].strftime('%Y%m%d') if 'date' in parsed_data[0] else datetime.now().strftime('%Y%m%d')
    if video_path:
        video_name = os.path.splitext(os.path.basename(video_path))[0]
        excel_filename=f"output/{video_name}_shots_timecode_thumbnails_{date_str}.xlsx"
    else:
        excel_filename=f"output/output_{date_str}.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shot Data"
    
    headers = ['file_type', 'machine', 'user', 'date', 'filename', 
               'location', 'frames', 'storage', 'producer', 'operator', 'job', 'notes']
    
    if video_path and 'start_timecode' in parsed_data[0]:
        headers.extend(['start_timecode', 'end_timecode', 'middle_timecode', 'middle_frame', 'thumbnail'])
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    #Use temp directory for thumbnail storage during processing
    temp_dir = tempfile.mkdtemp()
    
    try:
        for row_idx, entry in enumerate(parsed_data, 2):
            col = 1
            ws.cell(row=row_idx, column=col, value=entry.get('file_type', '')); col += 1
            ws.cell(row=row_idx, column=col, value=entry.get('machine', '')); col += 1
            ws.cell(row=row_idx, column=col, value=entry.get('file_user', entry.get('user', ''))); col += 1
            
            date_val = entry.get('file_date', entry.get('date', ''))
            if isinstance(date_val, datetime):
                date_val = date_val.strftime('%Y-%m-%d')
            ws.cell(row=row_idx, column=col, value=date_val); col += 1
            
            ws.cell(row=row_idx, column=col, value=entry.get('filename', '')); col += 1
            ws.cell(row=row_idx, column=col, value=entry.get('location', '')); col += 1
            
            frames_val = entry.get('frames', entry.get('original_frame_range', ''))
            if isinstance(frames_val, list):
                frames_val = ','.join(sort_frame_ranges_numeric(frames_val))
            ws.cell(row=row_idx, column=col, value=frames_val); col += 1
            
            ws.cell(row=row_idx, column=col, value=entry.get('storage', '')); col += 1
            ws.cell(row=row_idx, column=col, value=entry.get('producer', '')); col += 1
            ws.cell(row=row_idx, column=col, value=entry.get('operator', '')); col += 1
            ws.cell(row=row_idx, column=col, value=entry.get('job', '')); col += 1
            ws.cell(row=row_idx, column=col, value=entry.get('notes', '')); col += 1
            
            if video_path and 'start_timecode' in entry:
                ws.cell(row=row_idx, column=col, value=entry.get('start_timecode', '')); col += 1
                ws.cell(row=row_idx, column=col, value=entry.get('end_timecode', '')); col += 1
                ws.cell(row=row_idx, column=col, value=entry.get('middle_timecode', '')); col += 1
                ws.cell(row=row_idx, column=col, value=entry.get('middle_frame', '')); col += 1
                
                if 'middle_frame' in entry:
                    thumbnail_path = os.path.join(temp_dir, f"thumb_{row_idx}.jpg")
                    video_info = get_video_info(video_path, args)
                    if video_info and create_thumbnail(video_path, entry['middle_frame'], thumbnail_path, video_info['fps']):
                        try:
                            img = xl_image.Image(thumbnail_path)
                            img.width = 96
                            img.height = 74
                            #Embed thumbnail in Excel cell
                            ws.add_image(img, f"{chr(64 + col)}{row_idx}")
                        except Exception as e:
                            print(f"Error embedding thumbnail: {e}")
                
                col += 1
    
    finally:
        #Clean up temp directory to avoid disk clutter
        shutil.rmtree(temp_dir, ignore_errors=True)
    
    #Auto-adjust column widths for readability
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    if video_path:
        for row in range(2, len(parsed_data) + 2):
            ws.row_dimensions[row].height = 80
    
    os.makedirs('output', exist_ok=True)
    wb.save(excel_filename)
    print(f"Excel output written to {excel_filename}")

def insert_to_mongodb(parsed_data, args):
    print("Inserting data into MongoDB...")
    try:
        client = MongoClient('mongodb://localhost:27017/')
        db = client['MediaPipelineDB']
        scripts_collection = db['script_runs']
        files_collection = db['file_data']
        
        current_user = getpass.getuser()
        submitted_date = datetime.now()
        
        #Avoid duplicate script run entries by grouping
        unique_files = {}
        for entry in parsed_data:
            key = (entry['machine'], entry['user'], entry['date'])
            unique_files[key] = entry
            
        for key, entry in unique_files.items():
            script_record = {
                'script_user': current_user,
                'machine': entry['machine'],
                'file_user': entry['user'],
                'file_date': entry['date'],
                'submitted_date': submitted_date
            }
            result1 = scripts_collection.insert_one(script_record)
            if args.verbose:
                print(f"Inserted script run record: {result1.inserted_id}")
        
        inserted_count = 0
        for entry in parsed_data:
            #Insert each frame range as a separate record
            for frame_range in entry['frames']:
                file_record = {
                    'file_user': entry['user'],
                    'file_date': entry['date'],
                    'location': entry['location'],
                    'frames': [frame_range],
                    'storage': entry['storage'],
                    'machine': entry['machine'],
                    'filename': entry['filename'],
                    'file_type': entry['file_type'],
                    'producer': entry.get('producer', ''),
                    'operator': entry.get('operator', ''),
                    'job': entry.get('job', ''),
                    'notes': entry.get('notes', '')
                }
                result2 = files_collection.insert_one(file_record)
                inserted_count += 1
                if args.verbose:
                    print(f"Inserted file data record: {result2.inserted_id}")
        
        print(f"Successfully inserted {inserted_count} records into MongoDB")
        client.close()
        
    except Exception as e:
        print(f"Error inserting to MongoDB: {e}")
        sys.exit(1)

def view_database():
    print("Viewing database contents...")
    try:
        client = MongoClient('mongodb://localhost:27017/')
        db = client['MediaPipelineDB']
        scripts_collection = db['script_runs']
        files_collection = db['file_data']
        
        print("=== DATABASE CONTENTS ===\n")
        
        print("SCRIPT RUNS COLLECTION:")
        print("-" * 50)
        scripts = scripts_collection.find()
        script_count = 0
        for i, script_run in enumerate(scripts, 1):
            script_count += 1
            print(f"{i}. Script User: {script_run.get('script_user')}")
            print(f"   Machine: {script_run.get('machine')}")
            print(f"   File User: {script_run.get('file_user')}")
            print(f"   File Date: {script_run.get('file_date')}")
            print(f"   Submitted: {script_run.get('submitted_date')}")
            print()
        
        print(f"Total script runs: {script_count}\n")
        
        print("FILE DATA COLLECTION:")
        print("-" * 50)
        files = files_collection.find()
        file_count = 0
        for i, file_data in enumerate(files, 1):
            file_count += 1
            print(f"{i}. User: {file_data.get('file_user')}")
            print(f"   Date: {file_data.get('file_date')}")
            print(f"   Location: {file_data.get('location')}")
            print(f"   Frames: {file_data.get('frames')}")
            print(f"   Storage: {file_data.get('storage')}")
            print(f"   Machine: {file_data.get('machine')}")
            print(f"   File Type: {file_data.get('file_type')}")
            print()
        
        print(f"Total file records: {file_count}")
        client.close()
        
    except Exception as e:
        print(f"Error viewing database: {e}")
        sys.exit(1)

def export_csv_by_date(date_str, args):
    print(f"Exporting CSV for date: {date_str}")
    try:
        client = MongoClient('mongodb://localhost:27017/')
        db = client['MediaPipelineDB']
        files_collection = db['file_data']
        try:
            target_date = datetime.strptime(date_str, '%Y%m%d')
        except ValueError:
            print(f"Error: Invalid date format {date_str}. Expected YYYYMMDD")
            sys.exit(1)
        
        cursor = files_collection.find({'file_date': target_date})
        db_data = list(cursor)
        if not db_data:
            print(f"No data found for date: {date_str}")
            client.close()
            return
        
        if args.verbose:
            print(f"Found {len(db_data)} records for date: {date_str}")
        
        output_dir = 'output'
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, f"frame-fixes-{date_str}.csv")
        
        location_groups = {}
        producer = operator = job = notes = ""
        for entry in db_data:
            location = entry['location']
            if location not in location_groups:
                location_groups[location] = []
            location_groups[location].extend(entry['frames'])
            if not producer and entry.get('producer'):
                producer = entry['producer']
            if not operator and entry.get('operator'):
                operator = entry['operator']
            if not job and entry.get('job'):
                job = entry['job']
            if not notes and entry.get('notes'):
                notes = entry['notes']
        
        #Look for Xytech file matching the date
        xytech_file_path = f"Xytech_{date_str}.txt"
        xytech_locations = []
        xytech_file = find_file(xytech_file_path)
        if xytech_file:
            try:
                xytech_data = parse_xytech_file(xytech_file)
                xytech_locations = xytech_data['locations']
                print(f"Found Xytech file: {xytech_file}")
            except Exception as e:
                print(f"Could not parse Xytech file: {e}")
                xytech_locations = []
        else:
            print(f"Xytech file {xytech_file_path} not found")
        
        output_entries = []
        if xytech_locations:
            for xytech_location in xytech_locations:
                if xytech_location in location_groups:
                    if location_groups[xytech_location]:
                        frame_ranges = sort_frame_ranges_numeric(location_groups[xytech_location])
                        for frame_range in frame_ranges:
                            output_entries.append([xytech_location, frame_range])
                    else:
                        output_entries.append([xytech_location, "No frames to fix"])
            for location in location_groups:
                if location not in xytech_locations:
                    print(f"Warning: Location {location} not in Xytech file, adding to end of file")
                    if location_groups[location]:
                        frame_ranges = sort_frame_ranges_numeric(location_groups[location])
                        for frame_range in frame_ranges:
                            output_entries.append([location, frame_range])
                    else:
                        output_entries.append([location, "No frames to fix"])
        else:
            for location in location_groups:
                if location_groups[location]:
                    frame_ranges = sort_frame_ranges_numeric(location_groups[location])
                    for frame_range in frame_ranges:
                        output_entries.append([location, frame_range])
                else:
                    output_entries.append([location, "No frames to fix"])
        
        try:
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow([f"Producer: {producer}", f"Operator: {operator}", f"Job: {job}", f"Notes: {notes}"])
                writer.writerow([])
                writer.writerow([])
                writer.writerow(["Location:", "Frames to Fix:"])
                for entry in output_entries:
                    writer.writerow(entry)
            print(f"Successfully created output file: {output_file}")
        except Exception as e:
            print(f"Error writing CSV file: {e}")
            sys.exit(1)
        client.close()
    except Exception as e:
        print(f"Error exporting CSV for date {date_str}: {e}")
        sys.exit(1)

def clear_database():
    print("Clearing the database...")
    try:
        client = MongoClient('mongodb://localhost:27017/')
        db = client['MediaPipelineDB']
        scripts_result = db['script_runs'].delete_many({})
        files_result = db['file_data'].delete_many({})
        print(f"Deleted {scripts_result.deleted_count} script run records")
        print(f"Deleted {files_result.deleted_count} file data records")
        print("Database cleared successfully")
        client.close()
    except Exception as e:
        print(f"Error clearing database: {e}")
        sys.exit(1)

def process_video_workflow(video_path, args):
    if not os.path.exists(video_path):
        print(f"Error: Video file {video_path} not found")
        return []
    
    print("Step 1: Analyzing video...")
    video_info = get_video_info(video_path, args)
    if not video_info:
        return []
    
    print("Step 2: Querying database for matching ranges...")
    matching_ranges = get_matching_ranges(video_info, args)
    if not matching_ranges:
        print("No matching ranges found in video duration")
        return []
    
    output_dir = "rendered_shots"
    os.makedirs(output_dir, exist_ok=True)
    
    print("Step 3: Processing matching ranges...")
    processed_ranges = []
    
    for i, range_data in enumerate(matching_ranges):
        if args.verbose:
            print(f"Processing range {i+1}/{len(matching_ranges)}: {range_data['original_frame_range']}")
        
        shot_filename = f"shot_{i+1:03d}_{range_data['clamped_start_frame']}-{range_data['clamped_end_frame']}.mp4"
        shot_path = os.path.join(output_dir, shot_filename)
        
        if render_shot_segment(video_path, range_data['clamped_start_frame'], 
                             range_data['clamped_end_frame'], shot_path, video_info['fps']):
            
            if args.verbose:
                print(f"Uploading {shot_filename} to Frame.io...")
            
            upload_success = upload_to_frameio(shot_path, range_data)
            range_data['frameio_uploaded'] = upload_success
            range_data['shot_path'] = shot_path
        
        processed_ranges.append(range_data)
    
    return processed_ranges

def main():
    args = parse_arguments()
    print("Starting script execution...")
    
    if args.clear:
        clear_database()
        return
    
    if args.date:
        export_csv_by_date(args.date, args)
        return
    
    if not (args.workFiles or args.query or args.view or args.process):
        print("No Baselight/Flame files selected, no query, view, or video processing requested")
        sys.exit(2)
    
    if args.process:
        if args.verbose:
            print(f"Processing video: {args.process}")
        
        processed_ranges = process_video_workflow(args.process, args)
        
        if processed_ranges and args.output == 'XLS':
            output_to_excel(processed_ranges, args.process, args)
        elif processed_ranges:
            print(f"Processed {len(processed_ranges)} video ranges")
            for range_data in processed_ranges:
                print(f"Range: {range_data['original_frame_range']} -> "
                      f"Timecode: {range_data['start_timecode']}-{range_data['end_timecode']}")
    
    if args.workFiles:
        print(f"Processing {len(args.workFiles)} work files with output: {args.output or 'none'}")
        parsed_data, xytech_data = parse_db_files(args)
        
        if args.verbose:
            print(f"Parsed {len(parsed_data)} entries")
            for entry in parsed_data:
                print(entry)
        
        if args.output == 'CSV':
            output_to_csv(parsed_data, xytech_data)
        elif args.output == 'DB':
            insert_to_mongodb(parsed_data, args)
        elif args.output == 'XLS':
            output_to_excel(parsed_data, args=args)
    
    if args.view:
        view_database()

if __name__ == "__main__":
    main()
